from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time

# open chrom
web = webdriver.Chrome()
web.get('https://www.google.com/travel/flights/search?tfs=CBwQAhoqEgoyMDI0LTEwLTAxKABqDAgDEggvbS8wNndqZnIMCAMSCC9tLzAxOTE0QAFIAXABggELCP___________wGYAQI&tfu=EgIIAyIA&hl=zh-cn&curr=CNY')
time.sleep(1)

# create excel
workbook = Workbook()
sheet = workbook.active  # 获取活动的工作表

# select
for i in range(1, 100):
    try:
        apath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[3]/ul/li[{}]/div/div[3]/div/div/button/div[3]'.format(i)
        button1 = web.find_element(By.XPATH, apath)
        button1.click()
        time.sleep(0.88)

        bpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[3]/ul/li[{}]/div/div[4]/div/div[1]'.format(i)
        btext = web.find_element(By.XPATH, bpath)
        sheet.cell(row=i + 1, column=1, value=btext.text)

    except Exception as e:
        break

# save
file_name = '航班信息上海-北京.xlsx'
workbook.save(file_name)

# for i in range(1, 100):
#     try:
#         apath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[3]/div/div/button/div[3]'.format(i)
#         button1 = web.find_element(By.XPATH, apath)
#         button1.click()
#         time.sleep(0.88)
#
#         bpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[4]'.format(i)
#         btext = web.find_element(By.XPATH, bpath)
#         sheet.cell(row=i + 1, column=1, value=btext.text)
#
#         cpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[5]'.format(i)
#         ctext = web.find_element(By.XPATH, cpath)
#         sheet.cell(row=i + 1, column=2, value=ctext.text)
#
#         dpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[6]'.format(i)
#         dtext = web.find_element(By.XPATH, dpath)
#         sheet.cell(row=i + 1, column=3, value=dtext.text)
#
#         epath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[12]/span[2]'.format(i)
#         etext = web.find_element(By.XPATH, epath)
#         sheet.cell(row=i + 1, column=4, value=etext.text)
#
#         fpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[12]/span[8]'.format((i))
#         ftext = web.find_element(By.XPATH, fpath)
#         sheet.cell(row=i + 1, column=5, value=ftext.text)
#
#         gpath = '//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/c-wiz/div[1]/c-wiz/div[2]/div[2]/div[4]/ul/li[{}]/div/div[4]/div/div[1]/div[12]/span[10]'.format(i)
#         gtext = web.find_element(By.XPATH, gpath)
#         sheet.cell(row=i + 1, column=6, value=gtext.text)
#
#     except Exception as e:
#         break